#!/usr/bin/env python3
"""
excel_toolkit.py — Advanced Excel operations beyond pandas.

Handles: multi-sheet workbooks, cell formatting, conditional formatting,
data validation, formulas, freeze panes, auto-filter, column widths,
sheet operations (copy, move, delete, rename), and workbook creation from scratch.

Usage:
    python3 excel_toolkit.py <operation> <input> [options]

Examples:
    python3 excel_toolkit.py sheets data.xlsx                          # List all sheets
    python3 excel_toolkit.py extract data.xlsx --sheet Sales -o sales.csv
    python3 excel_toolkit.py combine *.csv -o combined.xlsx            # Multiple CSVs -> sheets
    python3 excel_toolkit.py format data.xlsx --header-style bold,blue --autowidth -o styled.xlsx
    python3 excel_toolkit.py freeze data.xlsx --at B2 -o frozen.xlsx
    python3 excel_toolkit.py autofilter data.xlsx -o filtered.xlsx
    python3 excel_toolkit.py validate data.xlsx --column Status --values "Open,Closed,Pending"
    python3 excel_toolkit.py protect data.xlsx --password secret -o protected.xlsx
    python3 excel_toolkit.py create --columns "Name,Age,Email" --rows 0 -o template.xlsx

Requirements:
    pip install openpyxl pandas
"""

import argparse
import os
import sys
from pathlib import Path

# --------------------------------------------------------------------------- #
# Dependency check
# --------------------------------------------------------------------------- #

def _check_deps():
    try:
        import openpyxl
        import pandas
        return openpyxl, pandas
    except ImportError as e:
        print(f"Error: {e}. Install with: pip install openpyxl pandas")
        sys.exit(1)


# --------------------------------------------------------------------------- #
# Operations
# --------------------------------------------------------------------------- #

def op_sheets(args):
    """List all sheets in a workbook."""
    openpyxl, _ = _check_deps()
    wb = openpyxl.load_workbook(args.input, read_only=True)

    print(f"=== {args.input} ===")
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        print(f"  [{i}] {name} ({rows:,} rows x {cols} cols)")
    wb.close()


def op_extract(args):
    """Extract a single sheet to CSV/JSON/Excel."""
    _, pd = _check_deps()

    sheet = args.sheet or 0
    df = pd.read_excel(args.input, sheet_name=sheet)
    print(f"Extracted sheet '{sheet}': {len(df):,} rows x {df.shape[1]} columns")

    if args.output:
        ext = Path(args.output).suffix.lower()
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        if ext == '.csv':
            df.to_csv(args.output, index=False)
        elif ext == '.json':
            df.to_json(args.output, orient='records', indent=2, force_ascii=False)
        elif ext in ('.xlsx', '.xls'):
            df.to_excel(args.output, index=False, engine='openpyxl')
        else:
            print(f"Error: unsupported output format '{ext}'")
            sys.exit(1)
        size = os.path.getsize(args.output)
        print(f"Output: {args.output} ({size:,} bytes)")
    else:
        print(df.to_string(max_colwidth=50))


def op_combine(args):
    """Combine multiple files into a single multi-sheet workbook."""
    openpyxl, pd = _check_deps()

    if not args.output:
        print("Error: -o/--output is required")
        sys.exit(1)

    inputs = args.inputs
    writer = pd.ExcelWriter(args.output, engine='openpyxl')

    for filepath in inputs:
        p = Path(filepath)
        if not p.exists():
            print(f"  Warning: {filepath} not found, skipping")
            continue

        ext = p.suffix.lower()
        sheet_name = p.stem[:31]  # Excel max sheet name = 31 chars

        if ext == '.csv':
            df = pd.read_csv(filepath)
        elif ext == '.tsv' or ext == '.txt':
            df = pd.read_csv(filepath, sep='\t')
        elif ext in ('.xlsx', '.xls'):
            df = pd.read_excel(filepath)
        elif ext == '.json':
            df = pd.read_json(filepath)
        else:
            print(f"  Warning: unsupported format {ext} for {filepath}, skipping")
            continue

        df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"  {p.name} -> sheet '{sheet_name}' ({len(df):,} rows)")

    writer.close()
    size = os.path.getsize(args.output)
    print(f"\nOutput: {args.output} ({size:,} bytes, {len(inputs)} sheets)")


def op_format(args):
    """Apply formatting to Excel workbook."""
    openpyxl, pd = _check_deps()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from copy import copy

    wb = openpyxl.load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active

    # Header styling
    if args.header_style:
        styles = [s.strip().lower() for s in args.header_style.split(',')]
        font_kwargs = {}
        fill = None
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        if 'bold' in styles:
            font_kwargs['bold'] = True
        if 'italic' in styles:
            font_kwargs['italic'] = True

        COLOR_MAP = {
            'blue': '4472C4', 'green': '70AD47', 'red': 'FF0000',
            'orange': 'ED7D31', 'purple': '7030A0', 'dark': '333333',
            'gray': 'A5A5A5', 'yellow': 'FFC000',
        }

        for s in styles:
            if s in COLOR_MAP:
                fill = PatternFill(start_color=COLOR_MAP[s], end_color=COLOR_MAP[s], fill_type='solid')
                font_kwargs['color'] = 'FFFFFF'

        font = Font(**font_kwargs) if font_kwargs else None

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            cell.alignment = alignment

        print(f"Applied header style: {', '.join(styles)}")

    # Auto-width
    if args.autowidth:
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)
        print("Applied auto-width to columns")

    # Zebra striping
    if args.zebra:
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = light_fill
        print("Applied zebra striping")

    # Number format
    if args.number_format:
        col_name, fmt = args.number_format.split(':', 1)
        # Find column index
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == col_name:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt
                print(f"Applied number format '{fmt}' to column '{col_name}'")
                break
        else:
            print(f"Warning: column '{col_name}' not found")

    # Save
    output = args.output or args.input
    wb.save(output)
    size = os.path.getsize(output)
    print(f"Output: {output} ({size:,} bytes)")


def op_freeze(args):
    """Freeze panes at specified cell."""
    openpyxl, _ = _check_deps()

    wb = openpyxl.load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active

    freeze_at = args.at or 'A2'
    ws.freeze_panes = freeze_at
    print(f"Frozen panes at {freeze_at}")

    output = args.output or args.input
    wb.save(output)
    print(f"Output: {output}")


def op_autofilter(args):
    """Add auto-filter to all columns."""
    openpyxl, _ = _check_deps()
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active

    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    print(f"Auto-filter applied: A1:{last_col}{ws.max_row}")

    output = args.output or args.input
    wb.save(output)
    print(f"Output: {output}")


def op_validate_cells(args):
    """Add data validation dropdown to a column."""
    openpyxl, _ = _check_deps()
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active

    if not args.column or not args.values:
        print("Error: --column and --values are required")
        sys.exit(1)

    # Find column letter
    col_letter = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == args.column:
            col_letter = get_column_letter(col_idx)
            break

    if not col_letter:
        print(f"Error: column '{args.column}' not found in headers")
        sys.exit(1)

    values_str = ','.join(v.strip() for v in args.values.split(','))
    dv = DataValidation(
        type='list',
        formula1=f'"{values_str}"',
        allow_blank=True
    )
    dv.error = f'Value must be one of: {values_str}'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = f'Select: {values_str}'
    dv.promptTitle = args.column

    cell_range = f'{col_letter}2:{col_letter}{ws.max_row}'
    dv.add(cell_range)
    ws.add_data_validation(dv)

    print(f"Added validation to {args.column} ({cell_range}): [{values_str}]")

    output = args.output or args.input
    wb.save(output)
    print(f"Output: {output}")


def op_protect(args):
    """Protect workbook/sheet with password."""
    openpyxl, _ = _check_deps()

    wb = openpyxl.load_workbook(args.input)

    if args.sheet:
        ws = wb[args.sheet]
        ws.protection.sheet = True
        ws.protection.password = args.password or ''
        print(f"Protected sheet '{args.sheet}'")
    else:
        for ws in wb.worksheets:
            ws.protection.sheet = True
            ws.protection.password = args.password or ''
        print(f"Protected all sheets ({len(wb.worksheets)})")

    output = args.output or args.input
    wb.save(output)
    print(f"Output: {output}")


def op_create(args):
    """Create a new Excel workbook from scratch."""
    openpyxl, _ = _check_deps()
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if not args.output:
        print("Error: -o/--output is required")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.sheet or 'Sheet1'

    # Add column headers
    if args.columns:
        cols = [c.strip() for c in args.columns.split(',')]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_text = Font(bold=True, color='FFFFFF')

        for i, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col_name)
            cell.font = header_text
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(i)].width = max(len(col_name) + 4, 12)

        print(f"Created {len(cols)} columns: {', '.join(cols)}")

    # Add sample rows
    rows = int(args.rows) if args.rows else 0
    if rows > 0:
        for r in range(2, rows + 2):
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c, value='')
        print(f"Added {rows} empty rows")

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto-filter
    if args.columns:
        last_col = get_column_letter(len(cols))
        ws.auto_filter.ref = f"A1:{last_col}1"

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    size = os.path.getsize(args.output)
    print(f"Output: {args.output} ({size:,} bytes)")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def build_parser():
    parser = argparse.ArgumentParser(
        description='Excel Toolkit — advanced Excel operations',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    subparsers = parser.add_subparsers(dest='operation', help='Operation')

    # Common
    def add_common(p, input_required=True):
        if input_required:
            p.add_argument('input', help='Input Excel file')
        p.add_argument('-o', '--output', help='Output file')
        p.add_argument('--sheet', help='Target sheet name')

    # sheets
    p = subparsers.add_parser('sheets', help='List sheets')
    add_common(p)

    # extract
    p = subparsers.add_parser('extract', help='Extract sheet to file')
    add_common(p)

    # combine
    p = subparsers.add_parser('combine', help='Combine files into multi-sheet xlsx')
    p.add_argument('inputs', nargs='+', help='Input files')
    p.add_argument('-o', '--output', required=True, help='Output xlsx file')

    # format
    p = subparsers.add_parser('format', help='Apply formatting')
    add_common(p)
    p.add_argument('--header-style', help='Header style: bold,blue,italic,green,red,...')
    p.add_argument('--autowidth', action='store_true', help='Auto-fit column widths')
    p.add_argument('--zebra', action='store_true', help='Add zebra striping')
    p.add_argument('--number-format', help='Format column: "ColName:0.00" or "Price:#,##0.00"')

    # freeze
    p = subparsers.add_parser('freeze', help='Freeze panes')
    add_common(p)
    p.add_argument('--at', default='A2', help='Cell to freeze at (def: A2)')

    # autofilter
    p = subparsers.add_parser('autofilter', help='Add auto-filter')
    add_common(p)

    # validate
    p = subparsers.add_parser('validate', help='Add dropdown validation')
    add_common(p)
    p.add_argument('--column', required=True, help='Column name')
    p.add_argument('--values', required=True, help='Allowed values (comma-sep)')

    # protect
    p = subparsers.add_parser('protect', help='Protect workbook')
    add_common(p)
    p.add_argument('--password', help='Protection password')

    # create
    p = subparsers.add_parser('create', help='Create new workbook')
    add_common(p, input_required=False)
    p.add_argument('--columns', required=True, help='Column headers (comma-sep)')
    p.add_argument('--rows', default='0', help='Number of empty rows')

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    if not args.operation:
        parser.print_help()
        sys.exit(0)

    ops = {
        'sheets': op_sheets,
        'extract': op_extract,
        'combine': op_combine,
        'format': op_format,
        'freeze': op_freeze,
        'autofilter': op_autofilter,
        'validate': op_validate_cells,
        'protect': op_protect,
        'create': op_create,
    }

    op_func = ops.get(args.operation)
    if op_func:
        try:
            op_func(args)
        except KeyboardInterrupt:
            print("\nAborted.")
            sys.exit(130)
        except FileNotFoundError as e:
            print(f"Error: file not found: {e}")
            sys.exit(1)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == '__main__':
    main()
